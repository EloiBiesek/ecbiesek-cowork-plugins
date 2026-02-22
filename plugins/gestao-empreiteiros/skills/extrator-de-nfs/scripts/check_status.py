#!/usr/bin/env python3
"""
Estado atual da extração NFSe — genérico (qualquer obra).

Mostra completude por empreiteiro, PDFs-imagem pendentes e totais.

Uso:
    python3 scripts/check_status.py [--obra-dir PATH]
    python3 scripts/check_status.py --json nfse_extracted.json
"""

import os, sys, json, argparse
from collections import defaultdict

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS_DIR)
import constants_nfse


def main():
    parser = argparse.ArgumentParser(description="Status NFSe")
    parser.add_argument("--obra-dir",
                        help="Caminho para a pasta da obra")
    parser.add_argument("--json",
                        help="JSON de extração a verificar (default: <obra>/.nfse-state/nfse_extracted.json)")
    args = parser.parse_args()

    # Inicializar obra
    obra_dir = args.obra_dir
    if not obra_dir:
        if constants_nfse.is_obra_dir(os.getcwd()):
            obra_dir = os.getcwd()
        else:
            obra_dir = constants_nfse.BASE_DIR
    constants_nfse.init_obra(obra_dir)

    if not args.json:
        args.json = os.path.join(constants_nfse.STATE_DIR, "nfse_extracted.json")

    print("=" * 70)
    print(f"ESTADO DA EXTRAÇÃO NFSe — {constants_nfse.OBRA_NOME}")
    print("=" * 70)

    # JSON de extração
    if not os.path.exists(args.json):
        print(f"\n  Arquivo JSON não encontrado: {args.json}")
        print("  Rode extract_all_nfse.py primeiro.")
        return

    with open(args.json, "r", encoding="utf-8") as f:
        data = json.load(f)

    records = data.get("records", [])
    stats = data.get("stats", {})

    print(f"\nTotal de registros no JSON: {len(records)}")
    print(f"  Extração OK: {stats.get('extracted_ok', '?')}")
    print(f"  Campos faltantes: {stats.get('missing_fields', '?')}")

    # Completude por empreiteiro
    print(f"\n{'='*70}")
    print("COMPLETUDE POR EMPREITEIRO")
    print(f"{'='*70}")
    print(f"  {'Empreiteiro':<40s} {'PDFs':>5s} {'OK':>4s} {'Pend':>5s} {'Img':>4s}")
    print(f"  {'-'*60}")

    by_empr = defaultdict(list)
    for r in records:
        by_empr[r.get("empreiteiro_num", "??")].append(r)

    total_pdfs = 0
    total_ok = 0
    total_pending = 0
    total_image = 0

    for num in sorted(by_empr.keys()):
        recs = by_empr[num]
        n_pdfs = len(recs)
        n_ok = sum(1 for r in recs
                   if r.get("nf") and r.get("valor_total") and r.get("competencia"))
        n_image = sum(1 for r in recs
                      if "sem texto" in (r.get("observacao") or "").lower())
        n_pending = n_pdfs - n_ok

        total_pdfs += n_pdfs
        total_ok += n_ok
        total_pending += n_pending
        total_image += n_image

        name = recs[0].get("empreiteiro", num)[:38]
        flag = " *" if n_image > 0 else ""
        print(f"  {name:<40s} {n_pdfs:>5d} {n_ok:>4d} {n_pending:>5d} {n_image:>4d}{flag}")

    print(f"  {'-'*60}")
    print(f"  {'TOTAL':<40s} {total_pdfs:>5d} {total_ok:>4d} {total_pending:>5d} {total_image:>4d}")
    if total_image > 0:
        print(f"\n  * = tem PDFs-imagem (podem precisar de OCR)")

    # Campos faltantes detalhados
    fields = ["nf", "razao_social", "cnpj_prestador", "cno",
              "competencia", "valor_total", "inss", "iss"]
    print(f"\n{'='*70}")
    print("PREENCHIMENTO DE CAMPOS")
    print(f"{'='*70}")
    for field in fields:
        count = sum(1 for r in records if r.get(field) is not None)
        pct = 100 * count / len(records) if records else 0
        bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
        print(f"  {field:<20s} {count:>4d}/{len(records):<4d} [{bar}] {pct:>5.1f}%")

    # Planilha
    xlsx_path = constants_nfse.XLSX_PATH
    aba = constants_nfse.ABA_NFSE
    print(f"\n{'='*70}")
    print("PLANILHA")
    print(f"{'='*70}")
    if os.path.exists(xlsx_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            if aba in wb.sheetnames:
                ws = wb[aba]
                rows_in_sheet = 0
                for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
                    val = row[0]
                    if val is not None and str(val).strip().upper() != "TOTAL":
                        rows_in_sheet += 1
                print(f"  Aba '{aba}': {rows_in_sheet} linhas de dados")
                diff = len(records) - rows_in_sheet
                if diff > 0:
                    print(f"  {diff} registros no JSON ainda não estão na planilha")
                elif diff == 0:
                    print(f"  Planilha sincronizada com o JSON")
                else:
                    print(f"  Planilha tem {abs(diff)} linhas a mais que o JSON")
            else:
                print(f"  Aba '{aba}' não encontrada")
            wb.close()
        except ImportError:
            print("  openpyxl não instalado — não foi possível verificar planilha")
    else:
        print(f"  Planilha não encontrada: {xlsx_path}")

    # Veredicto
    print(f"\n{'='*70}")
    action_needed = []
    if total_pending > 0:
        action_needed.append(f"{total_pending} registro(s) com campos faltantes")
    if total_image > 0:
        action_needed.append(f"{total_image} PDF(s)-imagem (podem precisar OCR)")

    if not action_needed:
        print("VEREDICTO: TUDO ATUALIZADO — nenhuma ação necessária.")
    else:
        print("VEREDICTO: AÇÃO NECESSÁRIA:")
        for a in action_needed:
            print(f"  - {a}")
        print("\nPRÓXIMOS PASSOS:")
        if total_image > 0:
            print("  1. python3 scripts/ocr_nfse.py --json nfse_extracted.json --batch-size 15")
        print("  2. python3 scripts/populate_xlsx.py")
        print("  3. python3 scripts/validate_extraction.py")
    print("=" * 70)


if __name__ == "__main__":
    main()
