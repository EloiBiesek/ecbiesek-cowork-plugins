#!/usr/bin/env python3
"""
Extract NFSe data from all contractor PDF folders — genérico (qualquer obra).

Walks each empreiteiro folder, enters NOTA FISCAL/ (or configured NF subfolders)
if present, or scans the root folder for PDFs (e.g. AQUILAIS case).

Usage:
    python3 extract_all_nfse.py [--obra-dir PATH] [--output PATH] [--xlsx PATH] [--incremental]

Defaults (após init_obra):
    --obra-dir  auto-detectado ou perguntado interativamente
    --output    <obra>/.nfse-state/nfse_extracted.json
    --xlsx      <obra>/CONTROLE GERAL DE NOTAS FISCAIS DE EMPREITEIROS.xlsx

Requires: pdfplumber, openpyxl (for --incremental)
"""
import pdfplumber
import re
import os
import json
import sys
import argparse

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS_DIR)
import constants_nfse


def parse_br_number(s):
    """Parse Brazilian number format: 320.000,00 -> 320000.00"""
    s = s.strip()
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def detect_service_type(full_text):
    """
    Detect whether the NFSe is for construction or surveillance/other services.
    This matters because surveillance services (AQUILAIS) have different rules:
    - INSS = 0 (not 11%) because owner-operated, no employee payroll
    - CNO is not applicable (no construction site)
    - ISS is typically 5% (not 2-3% Simples construction)
    """
    vigilancia_patterns = [
        r'VIGIL[AÂ]NCIA', r'SEGURAN[CÇ]A', r'MONITORAMENTO',
        r'11\.02', r'RONDA',
    ]
    for pat in vigilancia_patterns:
        if re.search(pat, full_text, re.I):
            return "vigilancia"
    return "construcao"


def extract_nfse(pdf_path):
    """
    Extract NFSe fields from a single PDF.
    Handles two Porto Velho municipal formats:
      1. Standard format: header/value lines separated (most contractors)
      2. "Nota Portovelhense" format: VALOR TOTAL DO SERVIÇO R$ xxx inline (JVB early notes)
    Also handles the AQUILAIS vigilância format with "Data Fato Gerador".
    """
    result = {
        "nf": None, "razao_social": None, "cnpj_prestador": None,
        "cno": None, "competencia": None, "valor_total": None,
        "inss": None, "iss": None, "observacao": "",
        "tipo_servico": None,
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages_text = []
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    pages_text.append(t)
            if not pages_text:
                result["observacao"] = "PDF sem texto extraível (possível imagem)"
                return result
            full = "\n".join(pages_text)
            lines = full.split('\n')

            # Detect service type early — affects validation rules
            result["tipo_servico"] = detect_service_type(full)

            # === NF Number ===
            # Strategy: try multiple patterns in priority order
            # 1. "Tipo de Recolhimento ... <NF>" line (standard format)
            # 2. "Nº da Nota Fiscal" followed by number (AQUILAIS format)
            # 3. "Número da Nota" section (Nota Portovelhense format)
            for i, line in enumerate(lines):
                if 'Tipo de Recolhimento' in line or 'Local de Recolhimento' in line:
                    m = re.search(r'(\d+)\s*$', line)
                    if m:
                        nf = int(m.group(1))
                        if nf < 100000:
                            result["nf"] = nf
                            break
                if 'Nº da Nota Fiscal' in line or 'N° da Nota Fiscal' in line:
                    # Check same line first (e.g., "Nº da Nota Fiscal\n91")
                    m = re.search(r'(\d+)\s*$', line)
                    if m:
                        nf = int(m.group(1))
                        if nf < 100000:
                            result["nf"] = nf
                            break
                    # Check next lines
                    for j in range(i + 1, min(i + 3, len(lines))):
                        m = re.search(r'(\d+)\s*$', lines[j])
                        if m:
                            nf = int(m.group(1))
                            if nf < 100000:
                                result["nf"] = nf
                                break
                    if result["nf"]:
                        break

            # Pattern for "Nota Portovelhense": "Número da Nota" as standalone field
            if not result["nf"]:
                m = re.search(r'N[úu]mero\s+da\s+Nota\s*\n\s*0*(\d+)', full, re.I)
                if m:
                    nf = int(m.group(1))
                    if nf < 100000:
                        result["nf"] = nf

            # Fallback: NF number from filename
            if not result["nf"]:
                fname = os.path.basename(pdf_path)
                m = re.search(r'(?:NFSE?|NF)\s*(\d+)', fname, re.I)
                if m:
                    result["nf"] = int(m.group(1))

            # === Prestador - Razão Social ===
            m = re.search(r'(?:Raz[ãa]o\s*Social|Nome/Raz[ãa]o)[:\s]*([^\n]+)', full, re.I)
            if m:
                val = m.group(1).strip()
                val = re.sub(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}.*', '', val).strip()
                if len(val) > 3:
                    result["razao_social"] = val

            # === CNPJ Prestador ===
            prestador_section = ""
            m_prest = re.search(r'PRESTADOR(.*?)TOMADOR', full, re.I | re.S)
            if m_prest:
                prestador_section = m_prest.group(1)
            cnpjs = re.findall(
                r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})',
                prestador_section if prestador_section else full
            )
            if cnpjs:
                result["cnpj_prestador"] = cnpjs[0]

            # === CNO ===
            # CNO is only relevant for construction services — vigilância notes won't have one
            m = re.search(r'(?:CNO|C\.N\.O)[:\s(]*([0-9][0-9.\-/]+)', full, re.I)
            if not m:
                # Fallback: busca direta pelo CNO da obra configurada
                cno_digits = re.sub(r'[^0-9]', '', constants_nfse.CNO)
                if len(cno_digits) >= 6:
                    prefix = cno_digits[:6]
                    pat = '.'.join(prefix[i:i+2] if i+2 <= len(prefix) else prefix[i:] for i in range(0, len(prefix), 2))
                    m = re.search(r'(' + re.escape(pat) + r'[0-9.\-/]*)', full)
            if m:
                result["cno"] = m.group(1).strip().rstrip(')')

            # === Competência ===
            # Priority hierarchy:
            # 1. "Competência" field (standard format, e.g., "08/2023")
            # 2. "Período" in discriminação
            # 3. "MES"/"MÊS"/"COMPETENCIA" in discriminação text
            # 4. "Data Fato Gerador" (AQUILAIS format — DD/MM/YYYY, use month)
            # 5. "referente" pattern
            # 6. Filename fallback (handled in main loop)
            m = re.search(r'Compet[êe]ncia[:\s]*(\d{1,2})[/\-](\d{4})', full, re.I)
            if not m:
                m = re.search(r'Per[íi]odo[:\s]*(\d{1,2})[/\-](\d{4})', full, re.I)
            if not m:
                m = re.search(r'(?:MES|COMP|MÊS|COMPETENCIA)\s*(\d{1,2})\s*[-/]\s*(\d{4})', full, re.I)
            if not m:
                # "Data Fato Gerador" format (AQUILAIS): DD/MM/YYYY
                # The date can be on the same line or the next line after the header
                for i, line in enumerate(lines):
                    if 'Data Fato Gerador' in line:
                        # Try same line first
                        m_dfg = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', line)
                        # Try next line if not found
                        if not m_dfg and i + 1 < len(lines):
                            m_dfg = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', lines[i + 1])
                        if m_dfg:
                            day, month_str, year_str = m_dfg.group(1), m_dfg.group(2), m_dfg.group(3)
                            month = int(month_str)
                            year = int(year_str)
                            if 1 <= month <= 12 and 2020 <= year <= 2030:
                                result["competencia"] = f"{month:02d}/{year}"
                        break
            if not m:
                m = re.search(r'referente.*?(\d{1,2})\s*[-/]\s*(\d{4})', full, re.I)
            if m and not result["competencia"]:
                month = int(m.group(1))
                year = int(m.group(2))
                if 1 <= month <= 12 and 2020 <= year <= 2030:
                    result["competencia"] = f"{month:02d}/{year}"

            # === VALOR TOTAL ===
            # Two format strategies:
            #   Format 1 (standard): "VALOR SERVIÇO ... ISS" header line, values on next line
            #   Format 2 (Nota Portovelhense): "VALOR TOTAL DO SERVIÇO R$ 130.398,00" inline
            for i, line in enumerate(lines):
                if 'VALOR SERVI' in line.upper() and 'ISS' in line.upper():
                    if i + 1 < len(lines):
                        val_line = lines[i + 1]
                        numbers = re.findall(r'[\d]+(?:\.[\d]+)*(?:,[\d]+)?', val_line)
                        if numbers:
                            v = parse_br_number(numbers[0])
                            if v and v > 0:
                                result["valor_total"] = round(v, 2)
                            if len(numbers) >= 2:
                                iss_val = parse_br_number(numbers[-1])
                                if iss_val and iss_val > 0:
                                    result["iss"] = round(iss_val, 2)
                    break

            # Format 2: "VALOR TOTAL DO SERVIÇO R$ xxx" on same line
            if not result["valor_total"]:
                m = re.search(
                    r'VALOR\s+TOTAL\s+(?:DO\s+)?SERVI[CÇ]O\s+R\$\s*([\d]+(?:\.[\d]+)*(?:,[\d]+)?)',
                    full, re.I
                )
                if m:
                    v = parse_br_number(m.group(1))
                    if v and v > 0:
                        result["valor_total"] = round(v, 2)

            # === ISS (if not already found) ===
            if not result["iss"]:
                for i, line in enumerate(lines):
                    if 'Valor do ISS' in line or 'Valor do ISSQN' in line:
                        # Multi-column header? (Portovelhense format)
                        # "Valor Deduções (R$) Base de Cálculo (R$) Alíquota (%) Valor do ISSQN (R$) ..."
                        # Values line: "0,00 130.398,00 2,00 2.607,96 0,00 0,00"
                        # ISS is the 4th value (index 3)
                        is_multi_col = ('Dedu' in line or 'Base de' in line)
                        if i + 1 < len(lines):
                            numbers = re.findall(r'[\d]+(?:\.[\d]+)*(?:,[\d]+)?', lines[i + 1])
                            if is_multi_col and len(numbers) >= 4:
                                v = parse_br_number(numbers[3])
                                if v and v > 0:
                                    result["iss"] = round(v, 2)
                                    break
                            elif numbers:
                                v = parse_br_number(numbers[0])
                                if v and v > 0:
                                    result["iss"] = round(v, 2)
                                    break
                        # Also check inline: "Valor do ISSQN (R$) ... 2.607,96"
                        if not is_multi_col:
                            numbers = re.findall(r'[\d]+(?:\.[\d]+)*,\d{2}', line)
                            if numbers:
                                v = parse_br_number(numbers[-1])
                                if v and v > 0:
                                    result["iss"] = round(v, 2)
                                    break

            # Also try: "Alíquota (%)... ISS (R$)..." value line pattern (AQUILAIS)
            if not result["iss"]:
                m = re.search(r'AL[ÍI]QUOTA.*?ISS', full, re.I)
                if m:
                    after = full[m.end():]
                    numbers = re.findall(r'[\d]+(?:\.[\d]+)*,\d{2}', after[:200])
                    if numbers:
                        v = parse_br_number(numbers[-1])
                        if v and v > 0:
                            result["iss"] = round(v, 2)

            # === INSS ===
            for i, line in enumerate(lines):
                if 'INSS (R$)' in line and 'IR (R$)' in line:
                    # Multi-column? "PIS (R$) COFINS (R$) INSS (R$) IR (R$)..."
                    # INSS is 3rd value (index 2)
                    is_multi_col = ('PIS' in line or 'COFINS' in line)
                    if i + 1 < len(lines):
                        val_line = lines[i + 1]
                        numbers = re.findall(r'[\d]+(?:\.[\d]+)*(?:,[\d]+)?', val_line)
                        if is_multi_col and len(numbers) >= 3:
                            v = parse_br_number(numbers[2])
                            if v is not None and v >= 0:
                                result["inss"] = round(v, 2)
                        elif numbers:
                            v = parse_br_number(numbers[0])
                            if v is not None and v >= 0:
                                result["inss"] = round(v, 2)
                    break

            # "Nota Portovelhense" INSS format: separate "INSS (R$)" section
            if result["inss"] is None:
                m = re.search(r'INSS\s*\(R\$\)\s*\n\s*([\d]+(?:\.[\d]+)*(?:,[\d]+)?)', full)
                if m:
                    v = parse_br_number(m.group(1))
                    if v is not None and v >= 0:
                        result["inss"] = round(v, 2)

            # Track substitution notes
            m_sub = re.search(r'substitui.*?N[Fº°]\s*[ºª]?\s*0*(\d+)', full, re.I)
            if m_sub:
                sub_note = f"Substitui NF {m_sub.group(1)}"
                if result["observacao"]:
                    result["observacao"] += "; "
                result["observacao"] += sub_note

    except Exception as e:
        result["observacao"] = f"Erro ao ler PDF: {str(e)[:100]}"

    return result


def comp_from_filename(fname):
    """Try to extract competência from filename as fallback."""
    m = re.search(r'(?:COMP\s*)?(\d{1,2})\s*[-]\s*(\d{4})', fname, re.I)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if 1 <= month <= 12 and 2020 <= year <= 2030:
            return f"{month:02d}/{year}"
    return None


def load_existing_records(xlsx_path):
    """Read the NFSe tab and build set of registered records."""
    from openpyxl import load_workbook

    if not os.path.exists(xlsx_path):
        return set(), set(), 0

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    aba = constants_nfse.ABA_NFSE
    if aba not in wb.sheetnames:
        wb.close()
        return set(), set(), 0

    ws = wb[aba]
    existing = set()
    filenames = set()
    row_count = 0

    for row in ws.iter_rows(min_row=5, max_col=10, values_only=True):
        empreiteiro = row[0]
        nf = row[1]
        comp = row[5]

        if empreiteiro is None:
            continue
        if str(empreiteiro).strip().upper() == "TOTAL":
            continue

        row_count += 1
        empreiteiro_num = str(empreiteiro).strip()[:2]
        nf_val = int(nf) if nf is not None else None
        comp_val = str(comp).strip() if comp else None

        if nf_val is not None and comp_val:
            existing.add((empreiteiro_num, nf_val, comp_val))

    wb.close()
    return existing, filenames, row_count


def pdf_matches_existing(record, existing_set):
    """Check if a PDF's extracted data matches an already-registered record."""
    num = record.get("empreiteiro_num")
    nf = record.get("nf")
    comp = record.get("competencia")

    if nf is not None and comp:
        if (num, nf, comp) in existing_set:
            return True

    if nf is not None:
        for (e_num, e_nf, e_comp) in existing_set:
            if e_num == num and e_nf == nf:
                return True

    if nf is None and comp and record.get("valor_total") is None:
        for (e_num, e_nf, e_comp) in existing_set:
            if e_num == num and e_comp == comp:
                return True

    return False


def find_nf_pdfs(empreiteiro_path, empreiteiro_dir):
    """
    Find NFSe PDFs for a given empreiteiro.

    Strategy:
    1. Check each configured NF subfolder (constants_nfse.NF_SUBFOLDERS) → walk recursively
    2. If none found, look for NF/NFSE PDFs directly in the empreiteiro folder
       (case of empreiteiro 11 AQUILAIS)
    """
    pdfs = []

    # Try configured NF subfolders
    for sub in constants_nfse.NF_SUBFOLDERS:
        nf_folder = os.path.join(empreiteiro_path, sub)
        if os.path.isdir(nf_folder):
            for root, dirs, files in os.walk(nf_folder):
                for f in files:
                    if f.lower().endswith('.pdf'):
                        pdfs.append(os.path.join(root, f))

    if not pdfs:
        # Look for NF/NFSE PDFs directly in the empreiteiro folder (not recursing into subfolders)
        for f in os.listdir(empreiteiro_path):
            fpath = os.path.join(empreiteiro_path, f)
            if os.path.isfile(fpath) and f.lower().endswith('.pdf'):
                # Only include files that look like NFs
                if re.search(r'(NF|NFSE|NFse|nfse|nota\s*fiscal)', f, re.I):
                    pdfs.append(fpath)

    return sorted(pdfs)


def main():
    parser = argparse.ArgumentParser(description='Extract NFSe data — genérico (qualquer obra)')
    parser.add_argument('--obra-dir',
                        help='Caminho para a pasta da obra (contém pastas de empreiteiros)')
    parser.add_argument('--base',
                        help='(Legado) Alias para --obra-dir')
    parser.add_argument('--output',
                        help='Output JSON file path (default: <obra>/.nfse-state/nfse_extracted.json)')
    parser.add_argument('--xlsx',
                        help='Path to NFSe spreadsheet (default: auto via constants_nfse)')
    parser.add_argument('--incremental', action='store_true',
                        help='Skip PDFs already registered in the spreadsheet')
    parser.add_argument('--retry-images', action='store_true',
                        help='In incremental mode, also retry image PDFs')
    args = parser.parse_args()

    # Resolver obra-dir (--obra-dir > --base > auto-detect)
    obra_dir = args.obra_dir or args.base
    if not obra_dir:
        # Auto-detect: se CWD é uma pasta de obra, use-a
        if constants_nfse.is_obra_dir(os.getcwd()):
            obra_dir = os.getcwd()
        else:
            # Fallback: BASE_DIR padrão (relativo ao script)
            obra_dir = constants_nfse.BASE_DIR

    constants_nfse.init_obra(obra_dir)

    BASE = constants_nfse.BASE_DIR
    OUTPUT = args.output or os.path.join(constants_nfse.STATE_DIR, "nfse_extracted.json")

    if not os.path.isdir(BASE):
        print(f"ERRO: Diretório base não encontrado: {BASE}")
        sys.exit(1)

    print(f"Obra: {constants_nfse.OBRA_NOME}")
    print(f"Base: {BASE}")
    print(f"Output: {OUTPUT}")
    print()

    # Load existing records if incremental
    xlsx_path = args.xlsx or constants_nfse.XLSX_PATH
    existing_set = set()
    existing_count = 0
    if args.incremental:
        print(f"Modo incremental: lendo planilha existente...")
        existing_set, _, existing_count = load_existing_records(xlsx_path)
        print(f"  {existing_count} registros encontrados na planilha")
        print(f"  {len(existing_set)} chaves únicas (empreiteiro+NF+comp)")
        print()

    # Find empreiteiro directories — prefer constants_nfse config, fallback to listdir
    if constants_nfse.EMPR_FOLDERS:
        contractors = sorted([
            constants_nfse.EMPR_FOLDERS[n]
            for n in constants_nfse.ALL_EMPR
            if os.path.isdir(os.path.join(BASE, constants_nfse.EMPR_FOLDERS[n]))
        ])
    else:
        contractors = sorted([
            d for d in os.listdir(BASE)
            if os.path.isdir(os.path.join(BASE, d)) and re.match(r'\d{2}\s', d)
        ])

    all_records = []
    stats = {
        "total_pdfs": 0, "extracted_ok": 0, "missing_fields": 0,
        "errors": 0, "skipped_existing": 0
    }

    for contractor_dir in contractors:
        cpath = os.path.join(BASE, contractor_dir)
        contractor_num = contractor_dir[:2]
        contractor_label = contractor_dir

        pdfs = find_nf_pdfs(cpath, contractor_dir)

        skipped = 0
        processed = 0

        for pdf_path in pdfs:
            stats["total_pdfs"] += 1
            fname = os.path.basename(pdf_path)

            data = extract_nfse(pdf_path)

            # Fallback: competência from filename
            if not data["competencia"]:
                data["competencia"] = comp_from_filename(fname)

            record = {
                "empreiteiro": contractor_label,
                "empreiteiro_num": contractor_num,
                "arquivo": fname,
                "arquivo_path": pdf_path,
                "nf": data["nf"],
                "razao_social": data["razao_social"],
                "cnpj_prestador": data["cnpj_prestador"],
                "cno": data["cno"],
                "competencia": data["competencia"],
                "valor_total": data["valor_total"],
                "inss": data["inss"],
                "iss": data["iss"],
                "observacao": data["observacao"],
                "tipo_servico": data.get("tipo_servico", "construcao"),
            }

            # Incremental: skip if already in spreadsheet
            if args.incremental and existing_set:
                is_image = 'sem texto' in (data.get("observacao") or "").lower()
                if pdf_matches_existing(record, existing_set):
                    if not (args.retry_images and is_image):
                        stats["skipped_existing"] += 1
                        skipped += 1
                        continue

            missing = []
            if not data["nf"]:
                missing.append("nf")
            if not data["valor_total"]:
                missing.append("valor_total")
            if not data["competencia"]:
                missing.append("competencia")
            if missing:
                stats["missing_fields"] += 1
                if record["observacao"]:
                    record["observacao"] += "; "
                record["observacao"] += f"Campos faltantes: {', '.join(missing)}"
            else:
                stats["extracted_ok"] += 1

            all_records.append(record)
            processed += 1

        status = f"{processed:3d} novos"
        if skipped > 0:
            status += f", {skipped:3d} já existentes"
        print(f"  {contractor_label[:45]:45s} | {len(pdfs):3d} PDFs | {status}")

    # Sort by contractor number, then competência
    def sort_key(r):
        comp = r.get("competencia") or "99/9999"
        parts = comp.split("/")
        try:
            return (r["empreiteiro_num"], int(parts[1]) * 100 + int(parts[0]))
        except (ValueError, IndexError):
            return (r["empreiteiro_num"], 999999)

    all_records.sort(key=sort_key)

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump({"stats": stats, "records": all_records}, f, ensure_ascii=False, indent=2)

    print(f"\n=== RESUMO ===")
    print(f"Total PDFs encontrados: {stats['total_pdfs']}")
    if args.incremental:
        print(f"Já registrados (pular): {stats['skipped_existing']}")
    print(f"Extração OK (novos):    {stats['extracted_ok']}")
    print(f"Campos faltantes:       {stats['missing_fields']}")
    print(f"Registros no JSON:      {len(all_records)}")
    print(f"Salvo em: {OUTPUT}")
    if args.incremental and len(all_records) == 0:
        print("\nNenhum PDF novo para processar. Planilha já está atualizada.")


if __name__ == "__main__":
    main()
