#!/usr/bin/env python3
"""
Populate NFSe tab with extracted NFSe data — genérico (qualquer obra).

Usage:
    python3 populate_xlsx.py [--obra-dir PATH] [--json PATH] [--xlsx PATH] [--append]

Modes:
    Default (full):   Clears the tab and writes all records from JSON.
    --append:         Finds the last data row, inserts new records BEFORE
                      the TOTAL row, and updates SUM formulas.

If the spreadsheet does not exist, it creates a new one with proper structure.

Defaults (após init_obra):
    --json  <obra>/.nfse-state/nfse_extracted.json
    --xlsx  <obra>/CONTROLE GERAL DE NOTAS FISCAIS DE EMPREITEIROS.xlsx

Requires: openpyxl
"""
import json
import os
import sys
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS_DIR)
import constants_nfse


def find_total_row(ws):
    """Find the TOTAL row in the worksheet. Returns row number or None."""
    for row in range(ws.max_row, 4, -1):
        val = ws[f'A{row}'].value
        if val and str(val).strip().upper() == "TOTAL":
            return row
    return None


def append_records(ws, records):
    """Append new records to existing data in the worksheet."""
    total_row = find_total_row(ws)

    if total_row is None:
        print("AVISO: Linha TOTAL não encontrada. Adicionando ao final.")
        insert_row = ws.max_row + 1
    else:
        insert_row = total_row

    # Styles
    data_font = Font(name="Arial", size=10)
    money_font = Font(name="Arial", size=10)
    obs_font = Font(name="Arial", size=9, italic=True, color="888888")
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    group_top_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='medium', color='1F3864'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    ws.insert_rows(insert_row, amount=len(records))

    prev_empreiteiro = None
    if insert_row > 5:
        prev_empreiteiro = ws[f'A{insert_row - 1}'].value

    for i, rec in enumerate(records):
        row = insert_row + i
        is_new_group = (rec["empreiteiro"] != prev_empreiteiro)
        prev_empreiteiro = rec["empreiteiro"]
        border = group_top_border if is_new_group else thin_border
        fill = even_fill if (i % 2 == 0) else PatternFill()

        ws[f"A{row}"] = rec["empreiteiro"]
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=is_new_group)
        ws[f"A{row}"].border = border
        ws[f"A{row}"].fill = fill

        ws[f"B{row}"] = rec["nf"]
        ws[f"B{row}"].font = data_font
        ws[f"B{row}"].alignment = Alignment(horizontal='center')
        ws[f"B{row}"].border = border
        ws[f"B{row}"].fill = fill

        ws[f"C{row}"] = rec["razao_social"]
        ws[f"C{row}"].font = data_font
        ws[f"C{row}"].border = border
        ws[f"C{row}"].fill = fill

        ws[f"D{row}"] = rec["cnpj_prestador"]
        ws[f"D{row}"].font = data_font
        ws[f"D{row}"].alignment = Alignment(horizontal='center')
        ws[f"D{row}"].border = border
        ws[f"D{row}"].fill = fill

        ws[f"E{row}"] = rec["cno"]
        ws[f"E{row}"].font = data_font
        ws[f"E{row}"].alignment = Alignment(horizontal='center')
        ws[f"E{row}"].border = border
        ws[f"E{row}"].fill = fill

        ws[f"F{row}"] = rec["competencia"]
        ws[f"F{row}"].font = data_font
        ws[f"F{row}"].alignment = Alignment(horizontal='center')
        ws[f"F{row}"].border = border
        ws[f"F{row}"].fill = fill

        ws[f"G{row}"] = rec["valor_total"]
        ws[f"G{row}"].font = money_font
        ws[f"G{row}"].number_format = '#,##0.00'
        ws[f"G{row}"].alignment = Alignment(horizontal='right')
        ws[f"G{row}"].border = border
        ws[f"G{row}"].fill = fill

        ws[f"H{row}"] = rec["inss"]
        ws[f"H{row}"].font = money_font
        ws[f"H{row}"].number_format = '#,##0.00'
        ws[f"H{row}"].alignment = Alignment(horizontal='right')
        ws[f"H{row}"].border = border
        ws[f"H{row}"].fill = fill

        ws[f"I{row}"] = rec["iss"]
        ws[f"I{row}"].font = money_font
        ws[f"I{row}"].number_format = '#,##0.00'
        ws[f"I{row}"].alignment = Alignment(horizontal='right')
        ws[f"I{row}"].border = border
        ws[f"I{row}"].fill = fill

        ws[f"J{row}"] = rec["observacao"] if rec["observacao"] else None
        ws[f"J{row}"].font = obs_font
        ws[f"J{row}"].border = border
        ws[f"J{row}"].fill = fill
        ws[f"J{row}"].alignment = Alignment(wrap_text=True)

    new_total_row = insert_row + len(records)
    if total_row is not None:
        for col in ['G', 'H', 'I']:
            ws[f'{col}{new_total_row}'].value = f'=SUM({col}5:{col}{new_total_row - 1})'

    ws.auto_filter.ref = f"A4:J{new_total_row - 1}"


def main():
    parser = argparse.ArgumentParser(description='Populate NFSe spreadsheet — genérico')
    parser.add_argument('--obra-dir',
                        help='Caminho para a pasta da obra')
    parser.add_argument('--json',
                        help='Input JSON file (default: <obra>/.nfse-state/nfse_extracted.json)')
    parser.add_argument('--xlsx',
                        help='Target Excel file (default: auto via constants_nfse)')
    parser.add_argument('--append', action='store_true',
                        help='Append new records instead of replacing all data')
    args = parser.parse_args()

    # Inicializar obra
    obra_dir = args.obra_dir
    if not obra_dir:
        if constants_nfse.is_obra_dir(os.getcwd()):
            obra_dir = os.getcwd()
        else:
            obra_dir = constants_nfse.BASE_DIR
    constants_nfse.init_obra(obra_dir)

    # Defaults pós-init
    if not args.json:
        args.json = os.path.join(constants_nfse.STATE_DIR, "nfse_extracted.json")
    if not args.xlsx:
        args.xlsx = constants_nfse.XLSX_PATH

    # Load data
    with open(args.json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    records = data["records"]

    if not records:
        print("Nenhum registro no JSON. Nada a fazer.")
        return

    aba = constants_nfse.ABA_NFSE

    # Load or create workbook
    if os.path.exists(args.xlsx):
        wb = load_workbook(args.xlsx)
    else:
        print(f"Planilha não encontrada. Criando nova: {args.xlsx}")
        wb = Workbook()
        ws_default = wb.active
        ws_default.title = aba

    if aba not in wb.sheetnames:
        wb.create_sheet(aba)

    ws = wb[aba]

    # === APPEND MODE ===
    if args.append:
        append_records(ws, records)
        wb.save(args.xlsx)
        print(f"Adicionados {len(records)} registros à planilha (modo append)")
        print(f"Arquivo: {args.xlsx}")
        return

    # === FULL MODE: Clear existing content ===
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.border = Border()
            cell.number_format = 'General'

    # === Styles ===
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
    data_font = Font(name="Arial", size=10)
    money_font = Font(name="Arial", size=10)
    obs_font = Font(name="Arial", size=9, italic=True, color="888888")
    total_font = Font(name="Arial", bold=True, size=10)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    group_top_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='medium', color='1F3864'),
        bottom=Side(style='thin', color='B0B0B0')
    )

    # === Row 3: Title ===
    ws.merge_cells('A3:J3')
    ws['A3'] = f"CONTROLE DE NOTAS FISCAIS DE SERVIÇO — {constants_nfse.OBRA_NOME.upper()}"
    ws['A3'].font = title_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # === Row 4: Headers ===
    headers = [
        ("A", "Empreiteiro"), ("B", "Nº NF"),
        ("C", "Razão Social do Prestador"), ("D", "CNPJ do Prestador"),
        ("E", "CNO da Obra"), ("F", "Competência"),
        ("G", "Valor Total (R$)"), ("H", "INSS (R$)"),
        ("I", "ISS (R$)"), ("J", "Observação"),
    ]
    for col_letter, name in headers:
        cell = ws[f"{col_letter}4"]
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Column widths
    widths = {"A": 38, "B": 10, "C": 42, "D": 22, "E": 20, "F": 14,
              "G": 18, "H": 16, "I": 14, "J": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 30

    # === Data rows starting at 5 ===
    row = 5
    prev_empreiteiro = None

    for i, rec in enumerate(records):
        is_new_group = (rec["empreiteiro"] != prev_empreiteiro)
        prev_empreiteiro = rec["empreiteiro"]
        border = group_top_border if is_new_group else thin_border
        fill = even_fill if (i % 2 == 0) else PatternFill()

        ws[f"A{row}"] = rec["empreiteiro"]
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=is_new_group)
        ws[f"A{row}"].border = border
        ws[f"A{row}"].fill = fill

        ws[f"B{row}"] = rec["nf"]
        ws[f"B{row}"].font = data_font
        ws[f"B{row}"].alignment = Alignment(horizontal='center')
        ws[f"B{row}"].border = border
        ws[f"B{row}"].fill = fill

        ws[f"C{row}"] = rec["razao_social"]
        ws[f"C{row}"].font = data_font
        ws[f"C{row}"].border = border
        ws[f"C{row}"].fill = fill

        ws[f"D{row}"] = rec["cnpj_prestador"]
        ws[f"D{row}"].font = data_font
        ws[f"D{row}"].alignment = Alignment(horizontal='center')
        ws[f"D{row}"].border = border
        ws[f"D{row}"].fill = fill

        ws[f"E{row}"] = rec["cno"]
        ws[f"E{row}"].font = data_font
        ws[f"E{row}"].alignment = Alignment(horizontal='center')
        ws[f"E{row}"].border = border
        ws[f"E{row}"].fill = fill

        ws[f"F{row}"] = rec["competencia"]
        ws[f"F{row}"].font = data_font
        ws[f"F{row}"].alignment = Alignment(horizontal='center')
        ws[f"F{row}"].border = border
        ws[f"F{row}"].fill = fill

        ws[f"G{row}"] = rec["valor_total"]
        ws[f"G{row}"].font = money_font
        ws[f"G{row}"].number_format = '#,##0.00'
        ws[f"G{row}"].alignment = Alignment(horizontal='right')
        ws[f"G{row}"].border = border
        ws[f"G{row}"].fill = fill

        ws[f"H{row}"] = rec["inss"]
        ws[f"H{row}"].font = money_font
        ws[f"H{row}"].number_format = '#,##0.00'
        ws[f"H{row}"].alignment = Alignment(horizontal='right')
        ws[f"H{row}"].border = border
        ws[f"H{row}"].fill = fill

        ws[f"I{row}"] = rec["iss"]
        ws[f"I{row}"].font = money_font
        ws[f"I{row}"].number_format = '#,##0.00'
        ws[f"I{row}"].alignment = Alignment(horizontal='right')
        ws[f"I{row}"].border = border
        ws[f"I{row}"].fill = fill

        ws[f"J{row}"] = rec["observacao"] if rec["observacao"] else None
        ws[f"J{row}"].font = obs_font
        ws[f"J{row}"].border = border
        ws[f"J{row}"].fill = fill
        ws[f"J{row}"].alignment = Alignment(wrap_text=True)

        row += 1

    # === TOTAL row ===
    total_row = row
    ws.merge_cells(f'A{total_row}:F{total_row}')
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'A{total_row}'].font = total_font
    ws[f'A{total_row}'].fill = total_fill
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
    double_border = Border(
        top=Side(style='double', color='1F3864'),
        bottom=Side(style='double', color='1F3864')
    )
    ws[f'A{total_row}'].border = double_border

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{total_row}'].fill = total_fill
        ws[f'{col}{total_row}'].border = double_border

    for col in ['G', 'H', 'I']:
        cell = ws[f'{col}{total_row}']
        cell.value = f'=SUM({col}5:{col}{total_row - 1})'
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = double_border

    ws[f'J{total_row}'].fill = total_fill
    ws[f'J{total_row}'].border = double_border

    # Freeze panes & auto filter
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f"A4:J{total_row - 1}"
    ws.print_title_rows = '3:4'

    # Save
    wb.save(args.xlsx)
    print(f"Planilha salva com {len(records)} registros (linhas 5-{total_row - 1})")
    print(f"Linha TOTAL: {total_row}")
    print(f"Arquivo: {args.xlsx}")


if __name__ == "__main__":
    main()
