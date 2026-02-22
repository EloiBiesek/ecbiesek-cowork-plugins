#!/usr/bin/env python3
"""
Estado atual da extração SEFIP — Lagoa Clube Resort.

Uso:
    python3 scripts/check_status.py
"""

import os, sys, json
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import constants


def main():
    print("=" * 70)
    print(f"ESTADO DA EXTRAÇÃO SEFIP — {constants.OBRA_NOME}")
    print("=" * 70)

    # Planilha
    if os.path.exists(constants.XLSX_PATH):
        print("\n── PLANILHA: Células preenchidas por empreiteiro ──")
        wb = openpyxl.load_workbook(constants.XLSX_PATH, data_only=True)
        if constants.ABA_ALOCACAO in wb.sheetnames:
            ws = wb[constants.ABA_ALOCACAO]
            for empr_num in constants.ALL_EMPR:
                row = constants.EMPR_ROW.get(empr_num)
                if row is None:
                    continue
                filled = 0
                for (y, m), col_letter in constants.MONTH_COL.items():
                    col_idx = openpyxl.utils.column_index_from_string(col_letter)
                    val = ws.cell(row=row, column=col_idx).value
                    if val is not None and val != 0:
                        filled += 1
                name = constants.EMPR_NAMES.get(empr_num, "?")
                empty = f"({constants.TOTAL_MONTHS - filled} vazias)" if filled < constants.TOTAL_MONTHS else "(completo)"
                print(f"  {empr_num:02d} {name:30s} {filled:2d}/{constants.TOTAL_MONTHS} {empty}")
        else:
            print("  Aba 'Alocação de colaboradores' não encontrada na planilha.")
        wb.close()
    else:
        print(f"\n  Planilha não encontrada: {constants.XLSX_PATH}")
        print("  Rode primeiro a skill nfse-extractor para criar a planilha.")

    # State files
    print(f"\n── ARQUIVOS DE ESTADO ({constants.STATE_DIR}) ──")
    for fname in ["extractions_text.json", "extractions_ocr.json",
                   "changes_log.json", "divergences.json",
                   "needs_ocr.json", "ocr_zeros.json"]:
        fpath = os.path.join(constants.STATE_DIR, fname)
        if os.path.exists(fpath):
            data = json.load(open(fpath, encoding="utf-8"))
            if isinstance(data, dict):
                total = sum(len(v) for v in data.values()) if all(isinstance(v, dict) for v in data.values()) else len(data)
            else:
                total = len(data)
            print(f"  ✓ {fname}: {total} registros")
        else:
            print(f"  ✗ {fname}: não existe")

    # Divergências
    pending_divs = 0
    div_path = os.path.join(constants.STATE_DIR, "divergences.json")
    if os.path.exists(div_path):
        divs = json.load(open(div_path, encoding="utf-8"))
        pending = [d for d in divs if not d.get("resolved")]
        resolved = [d for d in divs if d.get("resolved")]
        pending_divs = len(pending)
        print(f"\n── DIVERGÊNCIAS: {pending_divs} pendentes, {len(resolved)} resolvidas ──")
        for d in pending:
            print(f"  Empr {d['empr']:02d} {d['month']}: planilha={d['planilha']} vs pdf={d['pdf']}")

    # Carregar ambos os states para cross-reference
    NON_SEFIP_KW = ['BOLETO FGTS', 'CRÉDITO INSS', 'CREDITO INSS',
                    'DCTFWEB', 'FOLHA DE PAGAMENTO', 'FOLHA PAGAMENTO',
                    'FOLHA DE PONTO', 'GUIA DO FGTS', 'HOLERITE',
                    'COMPROVANTE DE DECLARAÇÃO', 'COMPROVANTE DE DECLARACAO',
                    'PROTOCOLO DE ENVIO', 'PARCELAMENTO',
                    'COMPENSAÇÃO INSS', 'COMPENSACAO INSS',
                    'RELATÓRIO ANALÍTICO DA GPS', 'RELATORIO ANALITICO DA GPS']

    all_states = {}
    for fname in ["extractions_text.json", "extractions_ocr.json"]:
        fpath = os.path.join(constants.STATE_DIR, fname)
        if os.path.exists(fpath):
            all_states[fname] = json.load(open(fpath, encoding="utf-8"))
        else:
            all_states[fname] = {}

    # Construir mapa de valores resolvidos (por qualquer método)
    resolved = set()
    for data in all_states.values():
        for empr_key, months in data.items():
            if not isinstance(months, dict):
                continue
            for month, info in months.items():
                if info.get("value") is not None:
                    resolved.add((empr_key, month))

    # Extrações com no_match — PDFs que existem mas não foram reconhecidos
    no_match_count = 0
    no_match_entries = []
    for fname, label in [("extractions_text.json", "texto"), ("extractions_ocr.json", "OCR")]:
        data = all_states.get(fname, {})
        for empr_key, months in data.items():
            if not isinstance(months, dict):
                continue
            for month, info in months.items():
                method = info.get("method", "")
                if info.get("value") is None:
                    # Pular se já resolvido por outro método (ex: texto falhou, OCR acertou)
                    if (empr_key, month) in resolved:
                        continue
                    path = info.get("path", "")
                    filename = os.path.basename(path).upper()
                    # Ignorar não-SEFIP conhecidos
                    if any(kw in filename for kw in NON_SEFIP_KW):
                        continue
                    if method in ("no_match", "empty_text", "ocr_no_match"):
                        # Evitar duplicatas (mesmo empr+month de texto e OCR)
                        dup_key = (empr_key, month)
                        if dup_key not in {(e["empr"], e["month"]) for e in no_match_entries}:
                            no_match_count += 1
                            no_match_entries.append({
                                "empr": empr_key, "month": month,
                                "method": method, "source": label,
                                "file": os.path.basename(path)
                            })

    if no_match_entries:
        print(f"\n── EXTRAÇÕES SEM RESULTADO (no_match): {no_match_count} ──")
        for e in sorted(no_match_entries, key=lambda x: (x['empr'], x['month'])):
            print(f"  Empr {e['empr']:>2s} {e['month']} [{e['method']}] {e['file']}")

    # OCR pendente — cruzar com extractions_ocr.json e filtrar não-SEFIP
    NON_SEFIP_KEYWORDS = ['BOLETO FGTS', 'FOLHA DE PAGAMENTO', 'FOLHA PAGAMENTO',
                          'CRÉDITO INSS', 'CREDITO INSS', 'DCTFWeb', 'DCTFWEB',
                          'GUIA DO FGTS']
    real_ocr_pending = 0
    ocr_path = os.path.join(constants.STATE_DIR, "needs_ocr.json")
    if os.path.exists(ocr_path):
        needs_raw = json.load(open(ocr_path, encoding="utf-8"))
        ext_ocr = {}
        ocr_state_path = os.path.join(constants.STATE_DIR, "extractions_ocr.json")
        if os.path.exists(ocr_state_path):
            ext_ocr = json.load(open(ocr_state_path, encoding="utf-8"))
        needs = []
        ja_processados = []
        non_sefip = []
        for n in needs_raw:
            empr_key = str(n['empr'])
            month = n['month']
            filename = os.path.basename(n.get('path', '')).upper()
            # Filtrar documentos que não são SEFIP
            if any(kw.upper() in filename for kw in NON_SEFIP_KEYWORDS):
                non_sefip.append(n)
                continue
            ocr_val = ext_ocr.get(empr_key, {}).get(month, {}).get('value')
            if ocr_val is not None:
                ja_processados.append(n)
            else:
                needs.append(n)
        real_ocr_pending = len(needs)
        info_parts = []
        if ja_processados:
            info_parts.append(f"{len(ja_processados)} já processados")
        if non_sefip:
            info_parts.append(f"{len(non_sefip)} não-SEFIP ignorados")
        info = f" ({', '.join(info_parts)})" if info_parts else ""
        print(f"\n── PDFs PENDENTES DE OCR: {real_ocr_pending}{info} ──")
        for n in needs:
            print(f"  Empr {n['empr']:02d} {n['month']}: ...{n['path'][-60:]}")
        if non_sefip:
            print(f"  Ignorados (não são SEFIP):")
            for n in non_sefip:
                print(f"    Empr {n['empr']:02d} {n['month']}: {os.path.basename(n['path'])}")
    else:
        real_ocr_pending = 0

    # Zeros suspeitos
    zero_path = os.path.join(constants.STATE_DIR, "ocr_zeros.json")
    zeros_count = 0
    if os.path.exists(zero_path):
        zeros = json.load(open(zero_path, encoding="utf-8"))
        zeros_count = len(zeros)
        print(f"\n── ZEROS OCR SUSPEITOS: {zeros_count} ──")
        for z in zeros:
            print(f"  Empr {z['empr']:02d} {z['month']}")

    # ── VEREDICTO FINAL ──
    # Determinar se há trabalho pendente real
    action_needed = []
    if no_match_count > 0:
        action_needed.append(f"{no_match_count} PDF(s) SEFIP com no_match (documento sem padrão reconhecido)")
    if real_ocr_pending > 0:
        action_needed.append(f"{real_ocr_pending} PDF(s) SEFIP pendente(s) de OCR")
    if pending_divs > 0:
        action_needed.append(f"{pending_divs} divergência(s) pendente(s)")
    if zeros_count > 0:
        action_needed.append(f"{zeros_count} zero(s) OCR suspeito(s)")

    print(f"\n{'='*70}")
    if not action_needed:
        print("VEREDICTO: ✅ TUDO ATUALIZADO — nenhuma ação necessária.")
        print("  Todos os PDFs SEFIP já foram extraídos e a planilha está em dia.")
        print("  Rode novamente apenas se novos PDFs forem adicionados às pastas.")
    else:
        print("VEREDICTO: ⚠️  AÇÃO NECESSÁRIA:")
        for a in action_needed:
            print(f"  - {a}")
        print("\nPRÓXIMOS PASSOS:")
        if no_match_count > 0 or real_ocr_pending > 0:
            print("  1. python3 scripts/extract_text.py --force  # re-extrai todos os PDFs")
            print("  2. python3 scripts/extract_ocr.py --from-pending  # OCR nos escaneados")
            print("  3. python3 scripts/update_planilha.py       # atualiza planilha")
        if pending_divs > 0:
            print("  4. python3 scripts/resolve_divergences.py --list")
    print("=" * 70)


if __name__ == "__main__":
    main()
