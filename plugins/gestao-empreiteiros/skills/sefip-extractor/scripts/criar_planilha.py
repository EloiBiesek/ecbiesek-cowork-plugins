"""
Criar planilha-mestre do zero para uma obra.

Gera a planilha com abas "Alocação de colaboradores" e "RESUMO NOVO"
baseado na configuração do obra.json.
"""

import os
import sys

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _generate_months(start, end):
    """Gera lista de (ano, mes) de start a end inclusive."""
    months = []
    y, m = start
    ey, em = end
    while (y, m) <= (ey, em):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


_MESES_PT = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}


def criar_planilha(config, output_path):
    """Cria planilha-mestre a partir da configuração."""
    wb = Workbook()

    nome_obra = config["nome_obra"]
    empreiteiros = config["empreiteiros"]
    mes_inicio = tuple(config["mes_inicio"])
    mes_fim = tuple(config["mes_fim"])
    months = _generate_months(mes_inicio, mes_fim)
    aba_alocacao = config.get("aba_alocacao", "Alocação de colaboradores")
    aba_resumo = config.get("aba_resumo", "RESUMO NOVO")
    row_start = config.get("linha_inicio_empr", 5)

    # --- Aba: Alocação de colaboradores ---
    ws = wb.active
    ws.title = aba_alocacao

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Linha 1: Título
    ws["A1"] = f"CONTROLE DE ALOCAÇÃO - {nome_obra.upper()}"
    ws["A1"].font = Font(bold=True, size=14)

    # Linha 2: Subtítulo
    ws["A2"] = "Quantidade de trabalhadores CAT 01 por empreiteiro/mês"
    ws["A2"].font = Font(size=10, italic=True)

    # Linha 3: vazia

    # Linha 4: Headers
    ws["A4"] = "Nº"
    ws["A4"].font = header_font_white
    ws["A4"].fill = header_fill
    ws["B4"] = "EMPREITEIRO"
    ws["B4"].font = header_font_white
    ws["B4"].fill = header_fill

    for i, (y, m) in enumerate(months):
        col = i + 3  # C=3
        cell = ws.cell(row=4, column=col)
        cell.value = f"{_MESES_PT[m]}/{y}"
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Larguras fixas
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40

    # Linhas dos empreiteiros
    for idx, e in enumerate(empreiteiros):
        row = row_start + idx
        ws.cell(row=row, column=1, value=e["num"]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=e["nome_completo"])
        for col_offset in range(len(months)):
            cell = ws.cell(row=row, column=3 + col_offset)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Linha de TOTAL (SUM)
    total_row = row_start + len(empreiteiros)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for col_offset in range(len(months)):
        col_letter = get_column_letter(3 + col_offset)
        first = row_start
        last = row_start + len(empreiteiros) - 1
        cell = ws.cell(row=total_row, column=3 + col_offset)
        cell.value = f"=SUM({col_letter}{first}:{col_letter}{last})"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # --- Aba: RESUMO NOVO ---
    ws2 = wb.create_sheet(aba_resumo)
    ws2["A1"] = f"CONTROLE DE EXTRAÇÕES SEFIP - {nome_obra.upper()}"
    ws2["A1"].font = Font(bold=True, size=14)

    # Salvar
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import json

    config_path = os.path.join(_SCRIPT_DIR, "..", "obra.json")
    if not os.path.exists(config_path):
        print("ERRO: obra.json nao encontrado. Rode configurar_obra.py primeiro.")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    base_dir = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", ".."))
    output = os.path.join(base_dir, config.get("planilha", "Controle.xlsx"))

    if os.path.exists(output):
        print(f"Planilha ja existe: {output}")
        sys.exit(0)

    criar_planilha(config, output)
    print(f"Planilha criada: {output}")
