#!/usr/bin/env python3
"""
Atualiza a aba "Alocação de colaboradores" — Lagoa Clube Resort.

Uso:
    python3 scripts/update_planilha.py
    python3 scripts/update_planilha.py --dry-run
"""

import os, sys, json, argparse
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import constants

AMARELO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ROSA = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
COMMENT_AUTHOR = f"SEFIP-Extractor-{constants.OBRA_NOME.replace(' ', '')}"


def load_extractions():
    """Carrega e mescla resultados de texto e OCR."""
    merged = {}

    text_path = os.path.join(constants.STATE_DIR, "extractions_text.json")
    if os.path.exists(text_path):
        data = json.load(open(text_path, encoding="utf-8"))
        for empr, months in data.items():
            merged.setdefault(empr, {}).update(months)

    ocr_path = os.path.join(constants.STATE_DIR, "extractions_ocr.json")
    if os.path.exists(ocr_path):
        data = json.load(open(ocr_path, encoding="utf-8"))
        for empr, months in data.items():
            for month_str, info in months.items():
                existing = merged.get(empr, {}).get(month_str)
                # Usar OCR se: mês não existe ainda, ou valor de texto é None e OCR tem valor
                if existing is None or (existing.get("value") is None and info.get("value") is not None):
                    merged.setdefault(empr, {})[month_str] = info

    return merged


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--skip-zero', action='store_true', default=True)
    parser.add_argument('--no-skip-zero', action='store_true')
    args = parser.parse_args()

    skip_zero = not args.no_skip_zero

    extractions = load_extractions()
    if not extractions:
        print("Nenhuma extração encontrada. Rode extract_text.py primeiro.")
        return

    if not os.path.exists(constants.XLSX_PATH):
        print(f"Planilha não encontrada: {constants.XLSX_PATH}")
        print("Rode primeiro a skill nfse-extractor para criar a planilha.")
        return

    wb = openpyxl.load_workbook(constants.XLSX_PATH)

    if constants.ABA_ALOCACAO not in wb.sheetnames:
        print(f"Aba '{constants.ABA_ALOCACAO}' não encontrada. Criando...")
        wb.create_sheet(constants.ABA_ALOCACAO)

    ws = wb[constants.ABA_ALOCACAO]

    changes = []
    divergences = []
    skipped_zero = []
    skipped_same = 0

    for empr_str, months_data in sorted(extractions.items(), key=lambda x: int(x[0])):
        empr_num = int(empr_str)
        row = constants.EMPR_ROW.get(empr_num)
        if row is None:
            continue

        empr_name = ws.cell(row=row, column=2).value or f"Empr {empr_num}"

        for month_str, info in sorted(months_data.items()):
            value = info.get("value")
            method = info.get("method", "")
            col_idx = constants.MONTH_COL_IDX.get(month_str)

            if col_idx is None or value is None:
                continue

            if skip_zero and value == 0 and "ocr" in method:
                skipped_zero.append({
                    "empr": empr_num, "month": month_str,
                    "method": method, "path": info.get("path", "")
                })
                continue

            current = ws.cell(row=row, column=col_idx).value
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            if current == value:
                skipped_same += 1
                continue

            if current is not None and current != 0 and current != value:
                divergences.append({
                    "empr": empr_num, "name": empr_name,
                    "month": month_str, "col": col_letter, "row": row,
                    "planilha": current, "pdf": value,
                    "method": method, "path": info.get("path", ""),
                    "resolved": False, "resolution": None
                })
                continue

            changes.append({
                "empr": empr_num, "name": empr_name,
                "month": month_str, "col": col_letter, "row": row,
                "old": current, "new": value, "method": method
            })

            if not args.dry_run:
                ws.cell(row=row, column=col_idx, value=value)

    # Sinalização visual
    if not args.dry_run:
        for d in divergences:
            cell = ws.cell(row=d['row'], column=openpyxl.utils.column_index_from_string(d['col']))
            cell.fill = AMARELO
            comment_text = (
                f"DIVERGÊNCIA — CONFERIR\n"
                f"Planilha: {d['planilha']}  |  PDF: {d['pdf']}\n"
                f"Procurar 'Qtd. Trabalhadores' ou CAT 01."
            )
            cell.comment = Comment(comment_text, COMMENT_AUTHOR)
            cell.comment.width = 400
            cell.comment.height = 120

        for z in skipped_zero:
            col_idx = constants.MONTH_COL_IDX.get(z['month'])
            z_row = constants.EMPR_ROW.get(z['empr'])
            if col_idx and z_row:
                cell = ws.cell(row=z_row, column=col_idx)
                cell.fill = ROSA
                comment_text = (
                    f"ZERO OCR — CONFERIR\n"
                    f"OCR retornou 0 trabalhadores (provável erro)."
                )
                cell.comment = Comment(comment_text, COMMENT_AUTHOR)
                cell.comment.width = 400
                cell.comment.height = 120

    # Salvar
    has_visual_marks = divergences or skipped_zero
    if not args.dry_run and (changes or has_visual_marks):
        wb.save(constants.XLSX_PATH)
        if changes:
            print(f"\n✓ {len(changes)} células atualizadas na planilha.")
        if divergences:
            print(f"✓ {len(divergences)} divergências sinalizadas em amarelo.")
        if skipped_zero:
            print(f"✓ {len(skipped_zero)} zeros OCR sinalizados em rosa.")
    elif args.dry_run:
        print(f"\n[DRY RUN] {len(changes)} células seriam atualizadas.")

    # Salvar state
    os.makedirs(constants.STATE_DIR, exist_ok=True)

    log_path = os.path.join(constants.STATE_DIR, "changes_log.json")
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(changes, f, indent=2, ensure_ascii=False)

    div_path = os.path.join(constants.STATE_DIR, "divergences.json")
    with open(div_path, "w", encoding="utf-8") as f:
        json.dump(divergences, f, indent=2, ensure_ascii=False)

    if skipped_zero:
        zero_path = os.path.join(constants.STATE_DIR, "ocr_zeros.json")
        with open(zero_path, "w", encoding="utf-8") as f:
            json.dump(skipped_zero, f, indent=2, ensure_ascii=False)

    # Relatório
    print(f"\n{'='*60}")
    print(f"RESUMO DA ATUALIZAÇÃO")
    print(f"{'='*60}")
    print(f"  Células atualizadas: {len(changes)}")
    print(f"  Já com mesmo valor:  {skipped_same}")
    print(f"  Divergências:        {len(divergences)}")
    print(f"  Zeros OCR ignorados: {len(skipped_zero)}")

    if changes:
        print(f"\n  ALTERAÇÕES:")
        for c in changes:
            print(f"    Empr {c['empr']:02d} | {c['month']} | {c['col']}{c['row']} | {c['old']} → {c['new']}")

    if divergences:
        print(f"\n  DIVERGÊNCIAS:")
        for d in divergences:
            print(f"    Empr {d['empr']:02d} | {d['month']} | {d['col']}{d['row']} | planilha={d['planilha']} vs pdf={d['pdf']}")


if __name__ == "__main__":
    main()
