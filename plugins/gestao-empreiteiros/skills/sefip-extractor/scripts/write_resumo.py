#!/usr/bin/env python3
"""
Cria/atualiza a aba RESUMO NOVO da planilha-mestre com:
  - Seção 1 (linhas 1-19):  ESTADO ATUAL — cobertura por empreiteiro (sobrescrito a cada execução)
  - Seção 2 (linhas 20+):   LOG DE EXECUÇÕES — uma linha adicionada por execução

Uso:
    python3 write_resumo.py --run-stats run_stats.json
    python3 write_resumo.py  # apenas atualiza estado atual, sem nova linha de log
"""

import os, sys, json, argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import constants

# ── Estilos ───────────────────────────────────────────────────────────────────

AZUL_ESCURO  = "1F4E79"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "BDD7EE"
CINZA_HEADER = "D6DCE4"
CINZA_PAR    = "F2F2F2"
VERDE_BG     = "E2EFDA"
AMARELO_BG   = "FFF2CC"
VERMELHO_BG  = "FCE4D6"
LARANJA_BG   = "FCE4D6"
BRANCO       = "FFFFFF"
LOG_NOVO_BG  = "EBF3FB"   # azul bem claro — linha mais recente do log

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_competencia(month_str):
    """'2025-09' → 'set/25'"""
    if not month_str:
        return "—"
    MESES = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
    try:
        y, m = month_str.split("-")
        return f"{MESES[int(m)-1]}/{y[2:]}"
    except Exception:
        return month_str

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ── Estado atual (lê dos arquivos de estado) ──────────────────────────────────

def build_estado_atual():
    text_data = {}
    ocr_data  = {}
    for fname, target in [("extractions_text.json", text_data),
                           ("extractions_ocr.json",  ocr_data)]:
        path = os.path.join(constants.STATE_DIR, fname)
        if os.path.exists(path):
            try:
                target.update(json.load(open(path, encoding="utf-8")))
            except Exception:
                pass

    rows = []
    for key in [str(i) for i in range(1, 13)]:
        t = text_data.get(key, {})
        o = ocr_data.get(key, {})
        all_m = {**t, **o}
        ok    = sorted(m for m, v in all_m.items() if v.get("value") is not None)
        fail  = [m for m, v in all_m.items() if v.get("value") is None]
        total_ok   = len(ok)
        total_fail = len(fail)
        comp_ini   = ok[0]  if ok else None
        comp_fim   = ok[-1] if ok else None
        cobertura  = round(total_ok / constants.TOTAL_MONTHS * 100, 1)
        rows.append({
            "nome":       constants.EMPR_NAMES_FULL.get(key, key),
            "comp_ini":   comp_ini,
            "comp_fim":   comp_fim,
            "total_ok":   total_ok,
            "total_fail": total_fail,
            "cobertura":  cobertura,
        })
    return rows


# ── Escrita da aba RESUMO NOVO ────────────────────────────────────────────────

# Linhas fixas da estrutura
ROW_TITLE       = 1
ROW_SUBTITLE    = 2
ROW_UPDATED     = 3
ROW_SPACER1     = 4
ROW_SEC1_HEADER = 5
ROW_COL_HEADERS = 6
ROW_EMPR_START  = 7
ROW_EMPR_END    = 7 + len(constants.ALL_EMPR) - 1
ROW_SPACER2     = 19
ROW_SEC2_HEADER = 20
ROW_LOG_HEADERS = 21
ROW_LOG_START   = 22   # linhas 22+ = log de execuções (dinâmico)

# Colunas da seção ESTADO ATUAL
COL_EMPR      = "A"  # Empreiteiro
COL_INI       = "B"  # Competência inicial
COL_FIM       = "C"  # Competência final
COL_OK        = "D"  # Meses extraídos
COL_FAIL      = "E"  # Meses com falha
COL_COB       = "F"  # Cobertura (%)
COL_UPD       = "G"  # Última atualização

# Colunas do LOG (mesmas letras, novo significado)
LOG_COL_DT    = "A"  # Data/hora
LOG_COL_EMPR  = "B"  # Empreiteiros processados (quantos)
LOG_COL_NTEX  = "C"  # Novos (texto)
LOG_COL_NOCR  = "D"  # Novos (OCR)
LOG_COL_SKIP  = "E"  # Pulados (já ok)
LOG_COL_ERR   = "F"  # Erros
LOG_COL_DIV   = "G"  # Divergências detectadas
LOG_COL_DRES  = "H"  # Divergências resolvidas
LOG_COL_COMP  = "I"  # Competência mais recente coberta (global)
LOG_COL_OBS   = "J"  # Observações


def write_resumo(run_stats=None):
    """
    run_stats: dict opcional com dados desta execução
    {
        "novos_texto": int, "novos_ocr": int, "pulados": int,
        "erros": int, "divergencias_detectadas": int,
        "divergencias_resolvidas": int, "comp_mais_recente": str,
        "empreiteiros_processados": int, "observacoes": str
    }
    """
    wb = load_workbook(constants.XLSX_PATH)

    if constants.ABA_RESUMO not in wb.sheetnames:
        ws = wb.create_sheet(constants.ABA_RESUMO)
        # Mover para terceira posição
        wb.move_sheet(ws, offset=2)
    else:
        ws = wb[constants.ABA_RESUMO]

    # ── Larguras de colunas ──────────────────────────────────────────────────
    widths = {"A": 42, "B": 16, "C": 16, "D": 14, "E": 14,
              "F": 12, "G": 22, "H": 14, "I": 18, "J": 40}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── LINHA 1: Título ──────────────────────────────────────────────────────
    c = ws["A1"]
    c.value = f"CONTROLE DE EXTRAÇÕES SEFIP — {constants.OBRA_NOME.upper()}"
    c.font  = font(bold=True, color=BRANCO, size=13)
    c.fill  = fill(AZUL_ESCURO)
    c.alignment = align("center", "center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:J1")

    # ── LINHA 2: Subtítulo ───────────────────────────────────────────────────
    c = ws["A2"]
    c.value = "Alocação de Colaboradores CAT 01 — Estado Atual e Log de Execuções"
    c.font  = font(italic=True, color=BRANCO, size=10)
    c.fill  = fill(AZUL_MEDIO)
    c.alignment = align("center", "center")
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")

    # ── LINHA 3: Última atualização ─────────────────────────────────────────
    c = ws["A3"]
    c.value = f"Atualizado em: {now_str}"
    c.font  = font(italic=True, color="595959", size=9)
    c.fill  = fill("EBF3FB")
    c.alignment = align("right", "center")
    ws.row_dimensions[3].height = 14
    ws.merge_cells("A3:J3")

    # ── LINHA 4: Espaçador ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── LINHA 5: Cabeçalho seção 1 ──────────────────────────────────────────
    c = ws["A5"]
    c.value = "▌ ESTADO ATUAL — Cobertura por Empreiteiro"
    c.font  = font(bold=True, color=BRANCO, size=10)
    c.fill  = fill(AZUL_MEDIO)
    c.alignment = align("left", "center")
    ws.row_dimensions[5].height = 18
    ws.merge_cells("A5:J5")

    # ── LINHA 6: Cabeçalhos de colunas ──────────────────────────────────────
    headers = [
        ("A6", "Empreiteiro"),
        ("B6", "Competência Inicial"),
        ("C6", "Competência Final"),
        ("D6", "Meses OK"),
        ("E6", "Meses c/ Falha"),
        ("F6", "Cobertura (%)"),
        ("G6", "Última Atualização"),
    ]
    for addr, label in headers:
        c = ws[addr]
        c.value     = label
        c.font      = font(bold=True, size=9)
        c.fill      = fill(CINZA_HEADER)
        c.alignment = align("center", "center")
        c.border    = thin_border()
    ws.row_dimensions[6].height = 16

    # ── LINHAS 7–18: Dados por empreiteiro ──────────────────────────────────
    estado = build_estado_atual()
    for i, row_data in enumerate(estado):
        row_num = ROW_EMPR_START + i
        bg      = CINZA_PAR if i % 2 == 0 else BRANCO

        # Cobertura → cor de fundo semáforo
        cob = row_data["cobertura"]
        if cob >= 60:
            cob_bg = VERDE_BG
        elif cob >= 25:
            cob_bg = AMARELO_BG
        else:
            cob_bg = VERMELHO_BG if row_data["total_ok"] > 0 else "F2F2F2"

        cells = [
            (f"A{row_num}", row_data["nome"],               bg,    "left"),
            (f"B{row_num}", fmt_competencia(row_data["comp_ini"]), bg, "center"),
            (f"C{row_num}", fmt_competencia(row_data["comp_fim"]), bg, "center"),
            (f"D{row_num}", row_data["total_ok"],            bg,    "center"),
            (f"E{row_num}", row_data["total_fail"],          bg,    "center"),
            (f"F{row_num}", f"{cob:.1f}%",                  cob_bg,"center"),
            (f"G{row_num}", now_str,                         bg,    "center"),
        ]
        for addr, val, bg_color, h_align in cells:
            c = ws[addr]
            c.value     = val
            c.font      = font(size=9)
            c.fill      = fill(bg_color)
            c.alignment = align(h_align, "center")
            c.border    = thin_border()
        ws.row_dimensions[row_num].height = 15

    # ── LINHA 19: Espaçador ──────────────────────────────────────────────────
    ws.row_dimensions[19].height = 8

    # ── LINHA 20: Cabeçalho seção 2 ─────────────────────────────────────────
    c = ws["A20"]
    c.value = "▌ LOG DE EXECUÇÕES — Histórico de Atualizações"
    c.font  = font(bold=True, color=BRANCO, size=10)
    c.fill  = fill(AZUL_MEDIO)
    c.alignment = align("left", "center")
    ws.row_dimensions[20].height = 18
    ws.merge_cells("A20:J20")

    # ── LINHA 21: Cabeçalhos do log ─────────────────────────────────────────
    log_headers = [
        ("A21", "Data / Hora"),
        ("B21", "Empreiteiros"),
        ("C21", "Novos Texto"),
        ("D21", "Novos OCR"),
        ("E21", "Pulados"),
        ("F21", "Erros"),
        ("G21", "Diverg. Detectadas"),
        ("H21", "Diverg. Resolvidas"),
        ("I21", "Comp. Mais Recente"),
        ("J21", "Observações"),
    ]
    for addr, label in log_headers:
        c = ws[addr]
        c.value     = label
        c.font      = font(bold=True, size=9)
        c.fill      = fill(CINZA_HEADER)
        c.alignment = align("center", "center", wrap=True)
        c.border    = thin_border()
    ws.row_dimensions[21].height = 28

    # ── Apagar fundo de linhas anteriores do log (garante limpeza visual) ───
    for row_num in range(ROW_LOG_START, ws.max_row + 1):
        for col in ["A","B","C","D","E","F","G","H","I","J"]:
            c = ws[f"{col}{row_num}"]
            if c.fill and c.fill.patternType == "solid":
                if c.fill.fgColor.rgb in (LOG_NOVO_BG.upper(), "FF" + LOG_NOVO_BG.upper()):
                    c.fill = fill(BRANCO)

    # ── Adicionar nova linha de log (se run_stats fornecido) ─────────────────
    if run_stats:
        # Encontrar próxima linha vazia a partir de ROW_LOG_START
        next_row = ROW_LOG_START
        while ws.cell(next_row, 1).value is not None:
            next_row += 1

        stats = run_stats
        comp_rec = fmt_competencia(stats.get("comp_mais_recente"))
        obs      = stats.get("observacoes", "")

        log_values = [
            now_str,
            stats.get("empreiteiros_processados", "—"),
            stats.get("novos_texto", 0),
            stats.get("novos_ocr",   0),
            stats.get("pulados",     0),
            stats.get("erros",       0),
            stats.get("divergencias_detectadas",  0),
            stats.get("divergencias_resolvidas",  0),
            comp_rec,
            obs,
        ]

        col_letters = ["A","B","C","D","E","F","G","H","I","J"]
        h_aligns    = ["center","center","center","center","center",
                       "center","center","center","center","left"]
        for col_l, val, h_a in zip(col_letters, log_values, h_aligns):
            c = ws[f"{col_l}{next_row}"]
            c.value     = val
            c.font      = font(size=9, bold=(col_l == "A"))
            c.fill      = fill(LOG_NOVO_BG)
            c.alignment = align(h_a, "center")
            c.border    = thin_border()
        ws.row_dimensions[next_row].height = 15
        print(f"  Nova linha de log adicionada → linha {next_row}")

    # ── Salvar ───────────────────────────────────────────────────────────────
    wb.save(constants.XLSX_PATH)
    print(f"  RESUMO NOVO atualizado em {constants.XLSX_PATH}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-stats", help="JSON com estatísticas desta execução")
    args = parser.parse_args()

    run_stats = None
    if args.run_stats and os.path.exists(args.run_stats):
        run_stats = json.load(open(args.run_stats, encoding="utf-8"))
    elif args.run_stats:
        try:
            run_stats = json.loads(args.run_stats)
        except Exception:
            pass

    write_resumo(run_stats)


if __name__ == "__main__":
    main()
