#!/usr/bin/env python3
"""
VALIDADOR DE ESTRUTURA DE OBRA
=====================================
Verifica se pastas de obra seguem a estrutura padrão ECBIESEK (14 seções numeradas).
Identifica: pastas faltando, vazias, com nomeação incorreta, documentos ausentes.

Uso:
    python validar_estrutura.py --obra-dir "C:/caminho/para/OBRA LAGUNAS"
    python validar_estrutura.py --pasta-mae "C:/caminho/para/PASTA MÃE ECBIESEK"
    python validar_estrutura.py --listar-obras
    python validar_estrutura.py --pasta-mae "..." --xlsx "relatorio.xlsx"
    python validar_estrutura.py --obra-dir "..." --template "meu_template.json"
"""

import os, sys, json, argparse, time
from datetime import datetime

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS_DIR)

import constants_validador as cv


# ── Helpers ──────────────────────────────────────────────────────────────────

def tempo_formatado(segundos):
    if segundos < 60:
        return f"{segundos:.0f}s"
    return f"{int(segundos // 60)}min {int(segundos % 60)}s"


# ── Validação de uma obra ────────────────────────────────────────────────────

def validar_obra(obra_path, template):
    """Valida uma pasta de obra contra o template.

    Returns:
        dict com resultados da validação:
        {
            "obra_name": str,
            "obra_path": str,
            "total_template": int,
            "total_required": int,
            "presentes": [{"pattern": str, "folder": str, "required": bool, "empty": bool, "file_count": int}],
            "faltando": [{"pattern": str, "required": bool}],
            "docs_faltando": [{"section": str, "pattern": str}],
            "extras": [str],
            "naming_issues": [{"folder": str, "expected_number": str, "issue": str}],
            "children_results": [{"parent": str, "expected": str, "found": bool}],
            "conformidade_pct": float,
        }
    """
    obra_name = os.path.basename(obra_path)
    subfolders = cv.list_subfolders(obra_path)
    structure = template.get("structure", [])

    result = {
        "obra_name": obra_name,
        "obra_path": obra_path,
        "total_template": len(structure),
        "total_required": sum(1 for s in structure if s.get("required", False)),
        "presentes": [],
        "faltando": [],
        "docs_faltando": [],
        "extras": [],
        "naming_issues": [],
        "children_results": [],
    }

    matched_folders = set()

    for section in structure:
        pattern = section["pattern"]
        required = section.get("required", False)
        expected_files = section.get("expected_files", [])
        expected_children = section.get("children", [])

        # Procurar pasta que corresponda ao pattern
        found_folder = None
        for folder in subfolders:
            if cv.match_pattern(folder, pattern):
                found_folder = folder
                matched_folders.add(folder)
                break

        if found_folder:
            folder_path = os.path.join(obra_path, found_folder)
            empty = cv.is_folder_empty(folder_path)
            file_count = cv.count_files_recursive(folder_path)

            result["presentes"].append({
                "pattern": pattern,
                "folder": found_folder,
                "required": required,
                "empty": empty,
                "file_count": file_count,
            })

            # Verificar arquivos obrigatórios
            if expected_files:
                files_in_section = cv.list_files(folder_path)
                # Também verificar em subpastas de 1 nível
                for sub in cv.list_subfolders(folder_path):
                    files_in_section.extend(cv.list_files(os.path.join(folder_path, sub)))

                for file_pattern in expected_files:
                    found_file = any(
                        cv.match_file_pattern(f, file_pattern)
                        for f in files_in_section
                    )
                    if not found_file:
                        result["docs_faltando"].append({
                            "section": found_folder,
                            "pattern": file_pattern,
                        })

            # Verificar subpastas esperadas
            if expected_children:
                child_folders = cv.list_subfolders(folder_path)
                for child_spec in expected_children:
                    child_pattern = child_spec["pattern"]
                    child_found = any(
                        cv.match_pattern(cf, child_pattern) for cf in child_folders
                    )
                    result["children_results"].append({
                        "parent": found_folder,
                        "expected": child_pattern,
                        "found": child_found,
                        "required": child_spec.get("required", False),
                    })
        else:
            result["faltando"].append({
                "pattern": pattern,
                "required": required,
            })

    # Pastas extras (não correspondem a nenhum pattern do template)
    for folder in subfolders:
        if folder not in matched_folders:
            result["extras"].append(folder)

    # Verificar problemas de nomeação
    for folder in subfolders:
        m = cv._OBRA_SECTION_RE.match(folder)
        if not m:
            # Pasta com número mas sem o padrão correto "NN - "
            m2 = cv.re.match(r"^(\d{1,2})\s*[-–]\s*(.+)$", folder)
            if m2:
                result["naming_issues"].append({
                    "folder": folder,
                    "expected_number": m2.group(1).zfill(2),
                    "issue": "Formato incorreto (esperado: 'NN - DESCRIÇÃO')",
                })

    # Calcular conformidade
    total = len(structure)
    presentes = len(result["presentes"])
    result["conformidade_pct"] = round((presentes / total * 100) if total > 0 else 0, 1)

    return result


# ── Impressão de relatório ───────────────────────────────────────────────────

def imprimir_relatorio_obra(r):
    """Imprime relatório de conformidade de uma obra no console."""
    print()
    print("=" * 70)
    print(f"  RELATORIO DE CONFORMIDADE")
    print(f"  {r['obra_name']}")
    print("=" * 70)

    pct = r["conformidade_pct"]
    presentes = len(r["presentes"])
    total = r["total_template"]
    print(f"\n  Conformidade: {pct:.0f}% ({presentes}/{total} secoes padrao encontradas)")

    # Presentes
    if r["presentes"]:
        print(f"\n  PRESENTES ({presentes}):")
        for p in r["presentes"]:
            status = ""
            if p["empty"]:
                status = "  [VAZIA]"
            elif p["file_count"] == 0:
                status = "  [SEM ARQUIVOS]"
            print(f"    [OK] {p['folder']}{status}")

    # Faltando
    if r["faltando"]:
        faltando_obr = [f for f in r["faltando"] if f["required"]]
        faltando_opt = [f for f in r["faltando"] if not f["required"]]

        if faltando_obr:
            print(f"\n  FALTANDO — OBRIGATORIAS ({len(faltando_obr)}):")
            for f in faltando_obr:
                print(f"    [!!] {f['pattern']}")

        if faltando_opt:
            print(f"\n  FALTANDO — OPCIONAIS ({len(faltando_opt)}):")
            for f in faltando_opt:
                print(f"    [--] {f['pattern']}")

    # Pastas vazias
    vazias = [p for p in r["presentes"] if p["empty"]]
    if vazias:
        print(f"\n  PASTAS VAZIAS ({len(vazias)}):")
        for v in vazias:
            req = "  [OBRIGATORIA]" if v["required"] else ""
            print(f"    [!!] {v['folder']}{req}")

    # Documentos faltando
    if r["docs_faltando"]:
        print(f"\n  DOCUMENTOS OBRIGATORIOS FALTANDO ({len(r['docs_faltando'])}):")
        for d in r["docs_faltando"]:
            print(f"    [!!] {d['section']}: falta {d['pattern']}")

    # Problemas de nomeação
    if r["naming_issues"]:
        print(f"\n  PROBLEMAS DE NOMEACAO ({len(r['naming_issues'])}):")
        for ni in r["naming_issues"]:
            print(f"    [!!] '{ni['folder']}' — {ni['issue']}")

    # Extras
    if r["extras"]:
        print(f"\n  PASTAS EXTRAS — fora do template ({len(r['extras'])}):")
        for e in r["extras"]:
            print(f"    [i]  {e}")

    print()
    print("=" * 70)


def imprimir_resumo_geral(resultados):
    """Imprime tabela resumo de todas as obras analisadas."""
    print()
    print("=" * 70)
    print(f"  RESUMO GERAL — {len(resultados)} OBRA(S) ANALISADA(S)")
    print("=" * 70)
    print()

    # Cabeçalho
    header = f"  {'Obra':<35} {'Conf.':<8} {'Faltando':<10} {'Vazias':<8} {'Docs':<6}"
    print(header)
    print("  " + "-" * 67)

    for r in resultados:
        nome = r["obra_name"]
        if len(nome) > 33:
            nome = nome[:30] + "..."

        pct = f"{r['conformidade_pct']:.0f}%"
        faltando_obr = len([f for f in r["faltando"] if f["required"]])
        faltando_total = len(r["faltando"])
        vazias = len([p for p in r["presentes"] if p["empty"]])
        docs = len(r["docs_faltando"])

        faltando_str = f"{faltando_obr}" if faltando_obr == faltando_total else f"{faltando_obr} ({faltando_total})"

        print(f"  {nome:<35} {pct:<8} {faltando_str:<10} {vazias:<8} {docs:<6}")

    # Média
    if resultados:
        media = sum(r["conformidade_pct"] for r in resultados) / len(resultados)
        print()
        print(f"  {'MEDIA GERAL':<35} {media:.0f}%")

    print()
    print("=" * 70)


# ── Exportação Excel ─────────────────────────────────────────────────────────

def exportar_xlsx(resultados, output_path):
    """Exporta resultados para planilha Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n  ERRO: openpyxl nao instalado. Execute: pip install openpyxl")
        return False

    wb = Workbook()

    # ── Aba Resumo ──
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    headers_resumo = [
        "Obra", "Caminho", "Conformidade %", "Secoes Presentes",
        "Secoes Faltando (Obrig.)", "Secoes Faltando (Total)",
        "Pastas Vazias", "Docs Faltando", "Pastas Extras",
    ]
    for col, header in enumerate(headers_resumo, 1):
        cell = ws_resumo.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Dados
    for row_idx, r in enumerate(resultados, 2):
        faltando_obr = len([f for f in r["faltando"] if f["required"]])
        values = [
            r["obra_name"],
            r["obra_path"],
            r["conformidade_pct"],
            len(r["presentes"]),
            faltando_obr,
            len(r["faltando"]),
            len([p for p in r["presentes"] if p["empty"]]),
            len(r["docs_faltando"]),
            len(r["extras"]),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_resumo.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

        # Colorir conformidade
        conf_cell = ws_resumo.cell(row=row_idx, column=3)
        if r["conformidade_pct"] >= 90:
            conf_cell.fill = green_fill
        elif r["conformidade_pct"] >= 70:
            conf_cell.fill = yellow_fill
        else:
            conf_cell.fill = red_fill

    # Ajustar larguras
    ws_resumo.column_dimensions["A"].width = 35
    ws_resumo.column_dimensions["B"].width = 60
    for col_letter in "CDEFGHI":
        ws_resumo.column_dimensions[col_letter].width = 18

    ws_resumo.freeze_panes = "A2"
    ws_resumo.auto_filter.ref = ws_resumo.dimensions

    # ── Aba Detalhes ──
    ws_det = wb.create_sheet("Detalhes")

    headers_det = [
        "Obra", "Secao (Template)", "Pasta Encontrada",
        "Status", "Obrigatoria", "Vazia", "Qtd Arquivos",
    ]
    for col, header in enumerate(headers_det, 1):
        cell = ws_det.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for r in resultados:
        # Presentes
        for p in r["presentes"]:
            status = "Vazia" if p["empty"] else "OK"
            values = [
                r["obra_name"], p["pattern"], p["folder"],
                status, "Sim" if p["required"] else "Nao",
                "Sim" if p["empty"] else "Nao", p["file_count"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws_det.cell(row=row, column=col, value=val)
                cell.border = thin_border
            # Colorir status
            status_cell = ws_det.cell(row=row, column=4)
            if status == "OK":
                status_cell.fill = green_fill
            else:
                status_cell.fill = yellow_fill
            row += 1

        # Faltando
        for f in r["faltando"]:
            values = [
                r["obra_name"], f["pattern"], "",
                "FALTANDO", "Sim" if f["required"] else "Nao", "", "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws_det.cell(row=row, column=col, value=val)
                cell.border = thin_border
            ws_det.cell(row=row, column=4).fill = red_fill
            row += 1

    ws_det.column_dimensions["A"].width = 35
    ws_det.column_dimensions["B"].width = 35
    ws_det.column_dimensions["C"].width = 45
    for col_letter in "DEFG":
        ws_det.column_dimensions[col_letter].width = 15

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions

    # ── Aba Docs Faltando ──
    if any(r["docs_faltando"] for r in resultados):
        ws_docs = wb.create_sheet("Docs Faltando")
        headers_docs = ["Obra", "Secao", "Documento Esperado"]
        for col, header in enumerate(headers_docs, 1):
            cell = ws_docs.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        row = 2
        for r in resultados:
            for d in r["docs_faltando"]:
                values = [r["obra_name"], d["section"], d["pattern"]]
                for col, val in enumerate(values, 1):
                    cell = ws_docs.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                row += 1

        ws_docs.column_dimensions["A"].width = 35
        ws_docs.column_dimensions["B"].width = 45
        ws_docs.column_dimensions["C"].width = 30
        ws_docs.freeze_panes = "A2"

    # Salvar
    wb.save(output_path)
    return True


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Valida estrutura de pastas de obra ECBIESEK",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python validar_estrutura.py --obra-dir "C:/PASTA MAE/OBRA LAGUNAS"
  python validar_estrutura.py --pasta-mae "C:/PASTA MAE ECBIESEK"
  python validar_estrutura.py --pasta-mae "C:/PASTA MAE" --xlsx relatorio.xlsx
  python validar_estrutura.py --listar-obras
"""
    )
    parser.add_argument("--obra-dir", type=str, default=None,
                        help="Caminho para uma pasta de obra especifica")
    parser.add_argument("--pasta-mae", type=str, default=None,
                        help="Caminho para a Pasta Mae (valida todas as obras)")
    parser.add_argument("--listar-obras", action="store_true",
                        help="Lista obras encontradas e sai")
    parser.add_argument("--template", type=str, default=None,
                        help="Caminho para template JSON customizado")
    parser.add_argument("--xlsx", type=str, default=None,
                        help="Exportar relatorio para Excel (.xlsx)")
    parser.add_argument("--json", type=str, default=None,
                        help="Exportar resultados para JSON")
    args = parser.parse_args()

    # Determinar diretório de busca
    search_dir = args.pasta_mae or args.obra_dir or os.getcwd()

    if not os.path.isdir(search_dir):
        print(f"\n  ERRO: Diretorio nao encontrado: {search_dir}")
        sys.exit(1)

    # Carregar template
    template = cv.load_template(args.template)

    # Modo: listar obras
    if args.listar_obras:
        obras = cv.descobrir_obras(search_dir)
        if not obras:
            print(f"\n  Nenhuma obra encontrada em: {search_dir}")
            print("  (Obras sao pastas com subpastas numeradas: '01 - DOCUMENTOS', '02 - DOCUMENTOS', ...)")
            sys.exit(0)

        print(f"\n  Obras encontradas ({len(obras)}):\n")
        for i, obra in enumerate(obras):
            print(f"    [{i+1}] {obra['name']}")
            print(f"        {obra['path']}")
            print(f"        {obra['sections']} secoes numeradas detectadas")
            print()
        sys.exit(0)

    # Descobrir obras
    if args.obra_dir:
        # Validar uma obra específica
        if not cv.is_obra_dir(args.obra_dir):
            print(f"\n  AVISO: '{args.obra_dir}' nao parece ser uma pasta de obra")
            print("  (Esperado: subpastas numeradas '01 - DOCUMENTOS', '02 - DOCUMENTOS', ...)")
            resp = input("  Continuar mesmo assim? [s/N]: ").strip().lower()
            if resp != "s":
                sys.exit(0)
        obras_paths = [os.path.normpath(args.obra_dir)]
    else:
        obras = cv.descobrir_obras(search_dir)
        if not obras:
            print(f"\n  Nenhuma obra encontrada em: {search_dir}")
            print("  Use --obra-dir para especificar uma pasta de obra diretamente.")
            sys.exit(1)
        obras_paths = [o["path"] for o in obras]

    # Executar validação
    inicio = time.time()

    print()
    print("=" * 70)
    print(f"  VALIDADOR DE ESTRUTURA DE OBRA — ECBIESEK")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"  Template: {os.path.basename(args.template or cv.DEFAULT_TEMPLATE_PATH)}")
    print(f"  Obras a validar: {len(obras_paths)}")
    print("=" * 70)

    resultados = []
    for obra_path in obras_paths:
        r = validar_obra(obra_path, template)
        resultados.append(r)
        imprimir_relatorio_obra(r)

    # Resumo geral (se mais de uma obra)
    if len(resultados) > 1:
        imprimir_resumo_geral(resultados)

    # Exportar Excel
    if args.xlsx:
        xlsx_path = os.path.normpath(args.xlsx)
        if exportar_xlsx(resultados, xlsx_path):
            print(f"\n  Relatorio Excel salvo em: {xlsx_path}")
        else:
            print(f"\n  ERRO ao salvar relatorio Excel.")

    # Exportar JSON
    if args.json:
        json_path = os.path.normpath(args.json)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)
        print(f"\n  Resultados JSON salvos em: {json_path}")

    total = time.time() - inicio
    print(f"\n  Tempo total: {tempo_formatado(total)}")
    print()


if __name__ == "__main__":
    main()
